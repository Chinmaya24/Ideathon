from flask import Flask, render_template, request, send_file, flash, redirect, url_for
import openpyxl
import os
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.utils import ImageReader
import qrcode


app = Flask(__name__)
app.secret_key = "secret123"  # needed for flash messages

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, "registrations.xlsx")


def save_to_excel(name, email, phone, institution):
    try:
        if not os.path.exists(EXCEL_FILE):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Name", "Email", "Phone", "Institution"])
            wb.save(EXCEL_FILE)

        wb = openpyxl.load_workbook(EXCEL_FILE)
    except Exception:  # if file is corrupt, recreate it
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Email", "Phone", "Institution"])
        wb.save(EXCEL_FILE)
        wb = openpyxl.load_workbook(EXCEL_FILE)

    ws = wb.active
    ws.append([name, email, phone, institution])
    wb.save(EXCEL_FILE)


def generate_pass(name, email, institution):
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # Title
    c.setFont("Helvetica-Bold", 20)
    c.drawCentredString(width / 2, height - 100, "🎟️ DSATM IDEATHON 2025 PASS")

    # Details
    c.setFont("Helvetica", 14)
    c.drawString(100, height - 180, f"Name: {name}")
    c.drawString(100, height - 210, f"Email: {email}")
    c.drawString(100, height - 240, f"Institution: {institution}")

    # Footer
    c.setFont("Helvetica-Oblique", 12)
    c.drawCentredString(width / 2, 100, "Show this pass at the venue for entry ✅")

    c.showPage()
    c.save()
    buffer.seek(0)
    return buffer


@app.route("/", methods=["GET", "POST"])
def home():
    if request.method == "POST":
        name = request.form["name"]
        email = request.form["email"]
        phone = request.form["phone"]
        institution = request.form["institution"]

        save_to_excel(name, email, phone, institution)

        # Generate pass
        pdf_buffer = generate_pass(name, email, institution)
        return send_file(pdf_buffer, as_attachment=True,
                         download_name=f"{name}_IdeathonPass.pdf",
                         mimetype="application/pdf")

    return render_template("index.html")

def generate_pass(name, email, institution):
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # Background
    c.setFillColorRGB(0.05, 0.28, 0.65)
    c.roundRect(40, height-400, width-80, 320, 20, fill=1, stroke=0)

    # Title
    c.setFont("Helvetica-Bold", 24)
    c.setFillColorRGB(1,1,1)
    c.drawCentredString(width / 2, height - 100, "🎟️ DSATM IDEATHON 2025")

    # Participant details
    c.setFont("Helvetica-Bold", 16)
    c.drawString(170, height - 180, f"Name: {name}")
    c.drawString(170, height - 210, f"Email: {email}")
    c.drawString(170, height - 240, f"Institution: {institution}")

    # Generate QR code
    qr_data = f"{name} - {institution}"
    qr = qrcode.QRCode(box_size=3, border=1)
    qr.add_data(qr_data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")

    # Convert QR code to ImageReader
    qr_buffer = io.BytesIO()
    img.save(qr_buffer, format='PNG')
    qr_buffer.seek(0)
    qr_image = ImageReader(qr_buffer)

    # Draw QR code
    c.drawImage(qr_image, width - 180, height - 300, width=120, height=120)

    # Footer
    c.setFont("Helvetica-Oblique", 12)
    c.drawCentredString(width / 2, height - 430, "Show this pass at the venue for entry ✅")

    c.showPage()
    c.save()
    buffer.seek(0)
    return buffer


if __name__ == "__main__":
    app.run(debug=True)
